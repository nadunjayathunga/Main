import warnings
import os
import pandas as pd
from data import company_info, table_info
from sqlalchemy import create_engine
from openpyxl import load_workbook
import sshtunnel

sshtunnel.SSH_TIMEOUT = 5.0
sshtunnel.TUNNEL_TIMEOUT = 5.0

warnings.simplefilter(action='ignore', category=FutureWarning)


PATH = r'C:\Masters\Masters'

PROMPT = '\nEnter company company Sr to proceed\nEnter "Q" to quit\n'

db_info = {'HOSTNAME': 'localhost',
           'USERNAME': 'nadunjayathunga',
           'PWDLOGIN': 'swis@123',
           'PWDDB': 'tech.123',
           'DB': 'nadunjayathunga$elite_security',
           'DBHOSTADDRESS': 'nadunjayathunga.mysql.pythonanywhere-services.com'}

ACTIVE = True
while ACTIVE:
    for cid in company_info:
        print(f'{cid["cid"]}\t{cid["data"]["long_name"]}')
    company_id = input(PROMPT)

    if company_id.lower() == 'q':
        print('Programme is quitting. Thank you')
        ACTIVE = False
    else:
        files_in_dir = []
        for folder, subfolers, filenames in os.walk(PATH):
            files_in_dir.append(filenames)

        if f'{company_info[int(company_id)-1]["data"]["long_name"]}.xlsx' not in files_in_dir[0]:
            print(
                f'{company_info[int(company_id)-1]["data"]["long_name"]}.xlsx not found in {PATH}')
            print('Please save all files\nProgramming is quitting. Thank you')
            ACTIVE = False

        else:
            work_book = load_workbook(
                f'{PATH}\{company_info[int(company_id)-1]["data"]["long_name"]}.xlsx',
                read_only=True)
            print('\nSr\tSheet name')
            for index, sheetname in enumerate(work_book.sheetnames):
                print(f'{index}\t{sheetname}')
                ACTIVE = False
            sheet_id = input('Please enter the Sr of the sheet to upload: ')
            sheet_name = work_book.sheetnames[int(sheet_id)]
            for index, value in enumerate(table_info):
                if table_info[index]["sheetname"] == sheet_name:
                    cid = index
                    break
            df_fdata = pd.read_excel(io=f'{PATH}\{company_info[int(company_id)-1]["data"]["long_name"]}.xlsx',
                                     sheet_name=sheet_name,
                                     usecols=table_info[index]['usecols'])
            # df_fdata.set_index(table_info[index]['index'], inplace=True)
            for i in company_info[int(company_id)-1]['data']:
                database = company_info[int(
                    company_id)-1]['data'].get('database')

            with sshtunnel.SSHTunnelForwarder(('ssh.pythonanywhere.com', 22), 
                                  ssh_username=db_info["USERNAME"], 
                                  ssh_password=db_info["PWDLOGIN"], 
                                  remote_bind_address=(db_info['DBHOSTADDRESS'], 3306)) as tunnel:

                engine = create_engine(f'mysql+pymysql://{db_info["USERNAME"]}:{db_info["PWDDB"]}@{db_info["HOSTNAME"]}:{tunnel.local_bind_port}/{db_info["DB"]}')

                df_fdata.to_sql(name=sheet_name, con=engine, if_exists='replace',index=False)

